#!/usr/bin/env python3
"""Scaffold reproducible data-derived figure and table artifacts for a deck workspace.

This is intentionally conservative. It does not try to "analyze" the deck for
the author. It finds simple tabular files, infers basic charts per file or
Excel worksheet, supports a small multi-series payload when columns align, and
writes a deterministic `assets/make_figures.py` script plus planning updates.
The generated script emits slide-ready figures, editable chart JSON, and compact
summary-table JSON so the main agent has concrete, rerunnable artifact paths to
refine. Metadata records both source provenance and figure export/readability
settings so later builds can audit whether generated charts were sized for
slide use.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Any


DATA_SUFFIXES = {".csv", ".tsv", ".xlsx", ".xls", ".json", ".jsonl", ".parquet", ".feather"}
COLUMNAR_SUFFIXES = {".parquet", ".feather"}
MAX_SERIES = 4
MAX_POINTS = 12
DEFAULT_FIGURE_SIZE_INCHES = [6.4, 3.6]
DEFAULT_FIGURE_DPI = 180
DEFAULT_TITLE_PT = 13
DEFAULT_AXIS_LABEL_PT = 8
DEFAULT_LEGEND_PT = 8
DEFAULT_X_LABEL_ROTATION = 35
ARTIFACT_REBUILD_CONTEXT_VERSION = "presentation_skill_artifact_rebuild_context_v1"


def _safe_id(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", value.strip()).strip("_").lower()
    return cleaned or "figure"


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, indent=2, ensure_ascii=False) + "\n"
    if path.exists():
        try:
            if path.read_text(encoding="utf-8") == text:
                return
        except OSError:
            pass
    path.write_text(text, encoding="utf-8")


def _resolve_path(workspace: Path, raw: str) -> Path:
    path = Path(raw).expanduser()
    if path.is_absolute():
        return path.resolve()
    cwd_candidate = (Path.cwd() / path).resolve()
    if cwd_candidate.exists():
        return cwd_candidate
    return (workspace / path).resolve()


def _relative_to_workspace(workspace: Path, path: Path) -> str:
    try:
        return str(path.resolve().relative_to(workspace.resolve()))
    except ValueError:
        return str(path.resolve())


def _file_fingerprint(path: Path) -> dict[str, Any]:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        size_bytes = path.stat().st_size
    except OSError:
        return {"source_sha256": "", "source_bytes": 0}
    return {"source_sha256": digest.hexdigest(), "source_bytes": size_bytes}


def _generated_artifact_paths(workspace: Path) -> set[Path]:
    manifest = _read_json(workspace / "assets" / "artifacts_manifest.json", {})
    paths: set[Path] = {
        (workspace / "assets" / "artifacts_manifest.json").resolve(),
        (workspace / "assets" / "analysis_summary.json").resolve(),
        (workspace / "assets" / "analysis_summary.md").resolve(),
        (workspace / "assets" / "attribution.csv").resolve(),
    }
    if not isinstance(manifest, dict):
        return paths
    for key in ("analysis_summary", "analysis_summary_markdown"):
        raw = str(manifest.get(key) or "").strip()
        if raw:
            paths.add(_resolve_path(workspace, raw))
    outputs = manifest.get("outputs")
    if not isinstance(outputs, list):
        return paths
    for output in outputs:
        if not isinstance(output, dict):
            continue
        for artifact in output.get("artifacts", []):
            if not isinstance(artifact, dict):
                continue
            raw = str(artifact.get("path") or "").strip()
            if raw:
                paths.add(_resolve_path(workspace, raw))
    return paths


def _candidate_data_files(workspace: Path, explicit_paths: list[str]) -> list[Path]:
    paths: list[Path] = []
    generated_paths = set() if explicit_paths else _generated_artifact_paths(workspace)
    staged_root = (workspace / "assets" / "staged").resolve()
    for raw in explicit_paths:
        path = _resolve_path(workspace, raw)
        if path.is_dir():
            for child in path.rglob("*"):
                if child.is_file() and child.suffix.lower() in DATA_SUFFIXES:
                    paths.append(child.resolve())
        else:
            paths.append(path.resolve())

    if not explicit_paths:
        for root in (
            workspace / "data",
            workspace / "assets" / "data",
            workspace / "assets" / "tables",
            workspace / "assets",
        ):
            if not root.exists():
                continue
            for path in root.rglob("*"):
                if path.is_file() and path.suffix.lower() in DATA_SUFFIXES:
                    paths.append(path.resolve())

    seen: set[Path] = set()
    unique: list[Path] = []
    for path in paths:
        resolved = path.resolve()
        if generated_paths and (
            resolved in generated_paths
            or staged_root in resolved.parents
        ):
            continue
        if resolved in seen:
            continue
        seen.add(resolved)
        unique.append(resolved)
    return unique[:24]


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def _read_csv_like(path: Path, delimiter: str) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        headers = [str(item or "").strip() for item in (reader.fieldnames or [])]
        rows = []
        for row in reader:
            rows.append({key: value for key, value in row.items() if key is not None})
            if len(rows) >= 200:
                break
    return headers, rows


def _worksheet_records(ws: Any) -> tuple[list[str], list[dict[str, Any]]]:
    row_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(row_iter)
    except StopIteration:
        return [], []
    headers = [str(item or "").strip() or f"col_{idx + 1}" for idx, item in enumerate(raw_headers)]
    rows: list[dict[str, Any]] = []
    for raw in row_iter:
        rows.append({headers[idx]: value for idx, value in enumerate(raw[: len(headers)])})
        if len(rows) >= 200:
            break
    return headers, rows


def _load_xlsx(path: Path) -> Any:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency path
        raise RuntimeError("openpyxl is required to inspect Excel files") from exc

    return load_workbook(path, read_only=True, data_only=True)


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = _load_xlsx(path)
    return _worksheet_records(wb.active)


def _iter_xlsx_tables(path: Path) -> list[dict[str, Any]]:
    wb = _load_xlsx(path)
    tables: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        headers, rows = _worksheet_records(ws)
        tables.append(
            {
                "headers": headers,
                "rows": rows,
                "sheet_name": ws.title,
                "source_label_suffix": f"::{ws.title}",
            }
        )
    return tables


def _read_json_table(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if path.suffix.lower() == ".jsonl":
        rows = []
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            item = json.loads(line)
            if isinstance(item, dict):
                rows.append(item)
            if len(rows) >= 200:
                break
    else:
        payload = json.loads(path.read_text(encoding="utf-8"))
        rows = payload if isinstance(payload, list) else payload.get("rows", []) if isinstance(payload, dict) else []
        rows = [row for row in rows if isinstance(row, dict)][:200]
    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(str(key))
    return headers, rows


def _read_columnar_table(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        import pandas as pd  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency path
        raise RuntimeError(f"pandas is required to inspect {path.suffix} files") from exc

    try:
        if path.suffix.lower() == ".parquet":
            frame = pd.read_parquet(path)
        else:
            frame = pd.read_feather(path)
    except Exception as exc:
        raise RuntimeError(
            f"pandas with pyarrow/fastparquet support is required to inspect {path.suffix} files: {exc}"
        ) from exc

    headers = [str(item or "").strip() or f"col_{idx + 1}" for idx, item in enumerate(frame.columns)]
    frame = frame.rename(columns={raw: header for raw, header in zip(frame.columns, headers)})
    frame = frame.head(200)
    frame = frame.where(pd.notnull(frame), None)
    rows = [
        {str(key): value for key, value in row.items()}
        for row in frame.to_dict(orient="records")
    ]
    return headers, rows


def _read_table(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_like(path, ",")
    if suffix == ".tsv":
        return _read_csv_like(path, "\t")
    if suffix in {".xlsx", ".xls"}:
        return _read_xlsx(path)
    if suffix in {".json", ".jsonl"}:
        return _read_json_table(path)
    if suffix in COLUMNAR_SUFFIXES:
        return _read_columnar_table(path)
    return [], []


def _table_sources(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        return _iter_xlsx_tables(path)
    headers, rows = _read_table(path)
    return [{"headers": headers, "rows": rows, "sheet_name": "", "source_label_suffix": ""}]


def _series_columns(headers: list[str], rows: list[dict[str, Any]]) -> list[str]:
    numeric_cols: list[str] = []
    for header in headers:
        values = [_parse_float(row.get(header)) for row in rows]
        present = [value for value in values if value is not None]
        if len(present) >= max(2, min(5, len(rows) // 3)):
            numeric_cols.append(header)
    return numeric_cols


def _chart_type_for(label_col: str, value_cols: list[str]) -> str:
    label_hint = label_col.lower()
    time_terms = ("time", "date", "day", "week", "month", "year", "cycle", "ct")
    if len(value_cols) > 1 and any(term in label_hint for term in time_terms):
        return "line"
    return "bar"


def _choose_label_and_value_columns(headers: list[str], numeric_cols: list[str]) -> tuple[str, list[str]]:
    numeric_set = set(numeric_cols)
    for header in headers:
        if header not in numeric_set:
            return header, numeric_cols[:MAX_SERIES]

    axis_terms = ("time", "date", "day", "week", "month", "year", "cycle")
    if len(numeric_cols) > 1:
        for header in headers:
            if header in numeric_set and any(term in header.lower() for term in axis_terms):
                values = [col for col in numeric_cols if col != header][:MAX_SERIES]
                if values:
                    return header, values

    value_cols = numeric_cols[:MAX_SERIES]
    label_col = headers[0] if headers else value_cols[0]
    if label_col in value_cols and len(value_cols) > 1:
        return label_col, [col for col in value_cols if col != label_col]
    return label_col, value_cols


def _infer_chart_spec(
    workspace: Path,
    path: Path,
    *,
    headers: list[str],
    rows: list[dict[str, Any]],
    sheet_name: str = "",
    source_label_suffix: str = "",
) -> dict[str, Any] | None:
    if not headers or not rows:
        return None

    numeric_cols = _series_columns(headers, rows)
    if not numeric_cols:
        return None

    label_col, value_cols = _choose_label_and_value_columns(headers, numeric_cols)
    value_col = value_cols[0]
    # Multi-series figures need a common x-axis. If fewer than two rows have
    # every selected numeric value, fall back to the strongest first series.
    complete_row_count = sum(
        1
        for row in rows
        if all(_parse_float(row.get(col)) is not None for col in value_cols)
    )
    if complete_row_count < 2:
        value_cols = [value_col]
    selected_columns = [label_col, *value_cols]
    id_parts = [path.stem]
    if sheet_name:
        id_parts.append(sheet_name)
    id_parts.append(value_col)
    chart_id = _safe_id("_".join(id_parts))[:56]
    title_prefix = (
        sheet_name.strip()
        if sheet_name
        else path.stem.replace("_", " ").replace("-", " ").title()
    )
    value_label = value_col if len(value_cols) == 1 else " + ".join(value_cols)
    title = f"{title_prefix}: {value_label}"
    source_path = _relative_to_workspace(workspace, path)
    return {
        "id": chart_id,
        "title": title,
        "source_path": source_path,
        "source_label": f"{source_path}{source_label_suffix}",
        "sheet_name": sheet_name,
        "label_col": label_col,
        "value_col": value_col,
        "value_cols": value_cols,
        "selected_columns": selected_columns,
        "chart_type": _chart_type_for(label_col, value_cols),
        "chart_json": f"assets/charts/{chart_id}.json",
        "figure_path": f"assets/figures/{chart_id}.png",
        "table_json": f"assets/tables/{chart_id}_summary.json",
        "target_variant": "image-sidebar",
        "target_box": "5.0x3.4 in",
        "figure_size_inches": DEFAULT_FIGURE_SIZE_INCHES,
        "figure_dpi": DEFAULT_FIGURE_DPI,
        "title_pt": DEFAULT_TITLE_PT,
        "axis_label_min_pt": DEFAULT_AXIS_LABEL_PT,
        "legend_pt": DEFAULT_LEGEND_PT,
        "x_label_rotation": DEFAULT_X_LABEL_ROTATION,
    }


def _infer_chart_specs(workspace: Path, path: Path) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    specs: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    for source in _table_sources(path):
        sheet_name = str(source.get("sheet_name") or "")
        label = f"{path}{source.get('source_label_suffix') or ''}"
        spec = _infer_chart_spec(
            workspace,
            path,
            headers=list(source.get("headers") or []),
            rows=list(source.get("rows") or []),
            sheet_name=sheet_name,
            source_label_suffix=str(source.get("source_label_suffix") or ""),
        )
        if spec is None:
            reason = "no usable numeric column found"
            if not source.get("headers") or not source.get("rows"):
                reason = "empty or headerless table"
            skipped.append({"path": str(label), "reason": reason})
            continue
        specs.append(spec)
    return specs, skipped


def _retarget_spec_id(spec: dict[str, Any], chart_id: str) -> None:
    spec["id"] = chart_id
    spec["chart_json"] = f"assets/charts/{chart_id}.json"
    spec["figure_path"] = f"assets/figures/{chart_id}.png"
    spec["table_json"] = f"assets/tables/{chart_id}_summary.json"


def _spec_identity(spec: dict[str, Any]) -> str:
    payload = {
        "source_path": spec.get("source_path", ""),
        "sheet_name": spec.get("sheet_name", ""),
        "label_col": spec.get("label_col", ""),
        "value_cols": spec.get("value_cols", []),
    }
    return json.dumps(payload, sort_keys=True, ensure_ascii=False)


def _disambiguate_spec_ids(specs: list[dict[str, Any]]) -> None:
    """Keep generated artifact paths unique without changing already-unique IDs."""
    base_counts: dict[str, int] = {}
    for spec in specs:
        base = _safe_id(str(spec.get("id") or "chart"))[:56]
        base_counts[base] = base_counts.get(base, 0) + 1

    used: set[str] = set()
    for spec in specs:
        base = _safe_id(str(spec.get("id") or "chart"))[:56]
        chart_id = base
        if base_counts.get(base, 0) > 1 or chart_id in used:
            digest = hashlib.sha1(_spec_identity(spec).encode("utf-8")).hexdigest()[:8]
            suffix = f"_{digest}"
            chart_id = f"{base[: max(1, 56 - len(suffix))]}{suffix}"
            counter = 2
            while chart_id in used:
                counter_suffix = f"_{digest}_{counter}"
                chart_id = f"{base[: max(1, 56 - len(counter_suffix))]}{counter_suffix}"
                counter += 1
        used.add(chart_id)
        if chart_id != spec.get("id"):
            _retarget_spec_id(spec, chart_id)


def _make_figures_script(specs: list[dict[str, Any]]) -> str:
    specs_literal = json.dumps(specs, indent=2, ensure_ascii=False)
    return f'''#!/usr/bin/env python3
"""Generate slide-ready figures for this deck workspace.

Generated by `scripts/scaffold_figure_artifacts.py`. Edit DATA_SPECS when you
need better grouping, filtering, labels, statistics, or chart types.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
from pathlib import Path
from typing import Any


WORKSPACE = Path(__file__).resolve().parents[1]
MPLCONFIGDIR = Path(os.environ.setdefault("MPLCONFIGDIR", str(WORKSPACE / "build" / "matplotlib-cache")))
MPLCONFIGDIR.mkdir(parents=True, exist_ok=True)

import matplotlib

matplotlib.use("Agg", force=True)
import matplotlib.pyplot as plt
try:
    from PIL import Image, ImageChops
except Exception:  # pragma: no cover - optional runtime dependency
    Image = None  # type: ignore[assignment]
    ImageChops = None  # type: ignore[assignment]

DATA_SPECS = {specs_literal}
REBUILD_CONTEXT_VERSION = "{ARTIFACT_REBUILD_CONTEXT_VERSION}"


def file_fingerprint(path: Path) -> dict[str, Any]:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return {{"source_sha256": digest.hexdigest(), "source_bytes": path.stat().st_size}}


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def read_csv_like(path: Path, delimiter: str) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


def read_xlsx(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(item or "").strip() or f"col_{{idx + 1}}" for idx, item in enumerate(next(rows))]
    except StopIteration:
        return []
    records = []
    for raw in rows:
        records.append({{headers[idx]: value for idx, value in enumerate(raw[: len(headers)])}})
    return records


def read_json_table(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".jsonl":
        records = []
        for line in path.read_text(encoding="utf-8").splitlines():
            if line.strip():
                item = json.loads(line)
                if isinstance(item, dict):
                    records.append(item)
        return records
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict) and isinstance(payload.get("rows"), list):
        return [item for item in payload["rows"] if isinstance(item, dict)]
    return []


def read_columnar_table(path: Path) -> list[dict[str, Any]]:
    import pandas as pd

    suffix = path.suffix.lower()
    if suffix == ".parquet":
        frame = pd.read_parquet(path)
    elif suffix == ".feather":
        frame = pd.read_feather(path)
    else:
        raise ValueError(f"Unsupported columnar data file: {{path}}")
    headers = [str(item or "").strip() or f"col_{{idx + 1}}" for idx, item in enumerate(frame.columns)]
    frame = frame.rename(columns={{raw: header for raw, header in zip(frame.columns, headers)}})
    frame = frame.where(pd.notnull(frame), None)
    return [
        {{str(key): value for key, value in row.items()}}
        for row in frame.to_dict(orient="records")
    ]


def read_records(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_like(path, ",")
    if suffix == ".tsv":
        return read_csv_like(path, "\\t")
    if suffix in {{".xlsx", ".xls"}}:
        return read_xlsx(path, sheet_name)
    if suffix in {{".json", ".jsonl"}}:
        return read_json_table(path)
    if suffix in {{".parquet", ".feather"}}:
        return read_columnar_table(path)
    raise ValueError(f"Unsupported data file: {{path}}")


def save_slide_figure(fig, path: Path, *, dpi: int = 180) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{{path.stem}}.tmp{{path.suffix}}")
    try:
        fig.savefig(
            tmp_path,
            dpi=dpi,
            bbox_inches="tight",
            pad_inches=0.04,
            facecolor="white",
            metadata={{"Software": "presentation-skill scaffold_figure_artifacts"}},
        )
    finally:
        plt.close(fig)
    if path.exists() and path.read_bytes() == tmp_path.read_bytes():
        tmp_path.unlink()
        return
    tmp_path.replace(path)


def corner_background_rgba(img: Any) -> tuple[int, int, int, int]:
    width, height = img.size
    corners = [
        img.getpixel((0, 0)),
        img.getpixel((max(0, width - 1), 0)),
        img.getpixel((0, max(0, height - 1))),
        img.getpixel((max(0, width - 1), max(0, height - 1))),
    ]
    channels = list(zip(*corners))
    return tuple(int(sorted(channel)[len(channel) // 2]) for channel in channels)  # type: ignore[return-value]


def image_exterior_whitespace(path: Path, *, tolerance: int = 12) -> dict[str, Any]:
    warning_threshold = 0.45
    if Image is None or ImageChops is None:
        return {{"checked": False, "reason": "Pillow unavailable", "warning_threshold": warning_threshold}}
    if not path.exists():
        return {{"checked": False, "reason": "image file missing", "warning_threshold": warning_threshold}}
    try:
        with Image.open(path) as raw:
            img = raw.convert("RGBA")
            width, height = img.size
            background = Image.new("RGBA", img.size, corner_background_rgba(img))
            diff = ImageChops.difference(img, background).convert("L")
            mask = diff.point(lambda value: 255 if value > tolerance else 0)
            bbox = mask.getbbox()
    except Exception as exc:
        return {{
            "checked": False,
            "reason": f"image whitespace check failed: {{exc}}",
            "warning_threshold": warning_threshold,
        }}
    if width <= 0 or height <= 0:
        return {{"checked": False, "reason": "image has invalid dimensions", "warning_threshold": warning_threshold}}
    if bbox is None:
        content_bbox = [0, 0, 0, 0]
        content_width = 0
        content_height = 0
    else:
        content_bbox = [int(item) for item in bbox]
        content_width = max(0, int(bbox[2] - bbox[0]))
        content_height = max(0, int(bbox[3] - bbox[1]))
    content_area = content_width * content_height
    total_area = max(1, int(width * height))
    exterior_fraction = max(0.0, min(1.0, 1.0 - (content_area / total_area)))
    return {{
        "checked": True,
        "width": int(width),
        "height": int(height),
        "content_bbox": content_bbox,
        "content_width": content_width,
        "content_height": content_height,
        "exterior_fraction": round(exterior_fraction, 4),
        "exterior_percent": round(exterior_fraction * 100, 1),
        "warning_threshold": warning_threshold,
        "high_exterior_whitespace": exterior_fraction >= warning_threshold,
    }}


def format_number(value: float | None) -> str:
    if value is None:
        return ""
    if abs(value) >= 100:
        return f"{{value:.0f}}"
    if abs(value) >= 10:
        return f"{{value:.1f}}"
    return f"{{value:.2f}}"


def summarize_series(labels: list[str], series_values: dict[str, list[float]]) -> dict[str, Any]:
    summaries: list[dict[str, Any]] = []
    for col, values in series_values.items():
        if not values:
            summaries.append(
                {{
                    "series": col,
                    "n": 0,
                    "min": None,
                    "max": None,
                    "mean": None,
                    "first": None,
                    "latest": None,
                    "delta": None,
                    "trend": "no plotted values",
                }}
            )
            continue
        first_value = values[0]
        latest_value = values[-1]
        delta = latest_value - first_value if len(values) > 1 else None
        if delta is None:
            trend = "single plotted point"
        elif delta > 0:
            trend = "increased"
        elif delta < 0:
            trend = "decreased"
        else:
            trend = "unchanged"
        summaries.append(
            {{
                "series": col,
                "n": len(values),
                "min": min(values),
                "max": max(values),
                "mean": sum(values) / len(values),
                "first": first_value,
                "latest": latest_value,
                "delta": delta,
                "trend": trend,
            }}
        )
    candidates = [item for item in summaries if item.get("n")]
    primary = ""
    if candidates:
        leader = max(
            candidates,
            key=lambda item: abs(float(item.get("delta") or item.get("latest") or 0)),
        )
        if leader.get("delta") is not None:
            primary = "{{}} {{}} from {{}} to {{}} across {{}} plotted points.".format(
                leader.get("series"),
                leader.get("trend"),
                format_number(leader.get("first")),
                format_number(leader.get("latest")),
                leader.get("n"),
            )
        else:
            primary = "{{}} latest value is {{}} from one plotted point.".format(
                leader.get("series"),
                format_number(leader.get("latest")),
            )
    return {{"primary": primary, "series": summaries}}


def chart_readability_warnings(
    labels: list[str],
    series_values: dict[str, list[float]],
    *,
    max_points: int,
    axis_label_min_pt: float,
) -> list[str]:
    warnings: list[str] = []
    if len(labels) >= max_points:
        warnings.append(
            "Figure uses the first {{}} complete rows; filter or split if omitted rows matter.".format(max_points)
        )
    long_label_count = len([label for label in labels if len(str(label)) > 18])
    if long_label_count:
        warnings.append(
            "{{}} x-axis label(s) exceed 18 characters; shorten labels or rotate/split before final delivery.".format(
                long_label_count
            )
        )
    if len(series_values) > 3:
        warnings.append(
            "{{}} series are plotted together; consider small multiples or a table for dense report slides.".format(
                len(series_values)
            )
        )
    if axis_label_min_pt < 8:
        warnings.append("Axis label size is below the 8 pt slide-readability floor.")
    return warnings


def write_json_if_changed(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, indent=2) + "\\n"
    if path.exists() and path.read_text(encoding="utf-8") == text:
        return
    path.write_text(text, encoding="utf-8")


def write_text_if_changed(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.read_text(encoding="utf-8") == text:
        return
    path.write_text(text, encoding="utf-8")


def artifact_fingerprint(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {{"sha256": "", "bytes": 0}}
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return {{"sha256": digest.hexdigest(), "bytes": path.stat().st_size}}


def data_specs_sha256() -> str:
    return hashlib.sha256(
        json.dumps(DATA_SPECS, sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()


def unique_strings(items: list[Any]) -> list[str]:
    seen: set[str] = set()
    values: list[str] = []
    for item in items:
        text = str(item or "").strip().replace("\\\\", "/")
        if not text or text in seen:
            continue
        seen.add(text)
        values.append(text)
    return values


def build_rebuild_context(
    outputs: list[dict[str, Any]],
    manifest_outputs: list[dict[str, Any]],
    *,
    producer_fingerprint: dict[str, Any],
    data_specs_hash: str,
) -> dict[str, Any]:
    source_paths = unique_strings(
        output.get("source_path") or (output.get("analysis_metadata") or {{}}).get("source_path")
        for output in outputs
    )
    figure_paths = unique_strings(output.get("figure_path") for output in outputs)
    chart_paths = unique_strings(output.get("chart_json") for output in outputs)
    table_paths = unique_strings(output.get("table_json") for output in outputs)
    artifact_paths: list[str] = []
    for output in manifest_outputs:
        for artifact in output.get("artifacts", []):
            if isinstance(artifact, dict):
                artifact_paths.append(str(artifact.get("path") or ""))
    return {{
        "context_version": REBUILD_CONTEXT_VERSION,
        "generated_by": "assets/make_figures.py",
        "producer_path": "assets/make_figures.py",
        "producer_sha256": producer_fingerprint.get("sha256", ""),
        "producer_bytes": producer_fingerprint.get("bytes", 0),
        "data_specs_sha256": data_specs_hash,
        "artifact_manifest": "assets/artifacts_manifest.json",
        "analysis_summary": "assets/analysis_summary.json",
        "analysis_summary_markdown": "assets/analysis_summary.md",
        "output_count": len(outputs),
        "source_count": len(source_paths),
        "source_paths": source_paths,
        "artifact_paths": unique_strings(artifact_paths),
        "outputs": {{
            "figures": figure_paths,
            "chart_json": chart_paths,
            "summary_tables": table_paths,
        }},
        "commands": {{
            "rebuild_figures": "python3 assets/make_figures.py",
            "inspect_manifest": "python3 scripts/inspect_artifact_manifest.py --workspace <deck_workspace> --manifest assets/artifacts_manifest.json",
            "auto_select_lead": "python3 scripts/apply_artifact_manifest_bindings.py --workspace <deck_workspace> --manifest assets/artifacts_manifest.json --auto-select --auto-select-mode lead",
            "auto_select_all": "python3 scripts/apply_artifact_manifest_bindings.py --workspace <deck_workspace> --manifest assets/artifacts_manifest.json --auto-select",
            "validate_planning": "python3 scripts/validate_planning.py --workspace <deck_workspace>",
        }},
        "command_cwd": {{
            "rebuild_figures": "deck_workspace",
            "inspect_manifest": "presentation-skill repository",
            "auto_select_lead": "presentation-skill repository",
            "auto_select_all": "presentation-skill repository",
            "validate_planning": "presentation-skill repository",
        }},
        "handoff_order": [
            "Run commands.rebuild_figures after editing DATA_SPECS or analysis code.",
            "Run commands.inspect_manifest to refresh alias, quality, and selection guidance.",
            "Run commands.auto_select_lead for a clean first pass or commands.auto_select_all for the full evidence triplet.",
            "Run commands.validate_planning before strict PPTX build.",
        ],
    }}


def build_chart(spec: dict[str, Any]) -> dict[str, Any]:
    data_path = WORKSPACE / spec["source_path"]
    rows = read_records(data_path, spec.get("sheet_name") or None)
    fingerprint = file_fingerprint(data_path)
    producer_fingerprint = artifact_fingerprint(Path(__file__))
    value_cols = [str(item) for item in (spec.get("value_cols") or [spec["value_col"]])]
    selected_columns = [
        str(item)
        for item in (spec.get("selected_columns") or [spec["label_col"], *value_cols])
        if str(item).strip()
    ]
    max_points = int(spec.get("max_points") or 12)
    raw_figure_size = spec.get("figure_size_inches") or [6.4, 3.6]
    figure_size_inches = (
        raw_figure_size
        if isinstance(raw_figure_size, list) and len(raw_figure_size) == 2
        else [6.4, 3.6]
    )
    figure_dpi = int(spec.get("figure_dpi") or 180)
    title_pt = float(spec.get("title_pt") or 13)
    axis_label_min_pt = float(spec.get("axis_label_min_pt") or 8)
    legend_pt = float(spec.get("legend_pt") or axis_label_min_pt)
    x_label_rotation = float(spec.get("x_label_rotation") or 35)
    labels: list[str] = []
    series_values: dict[str, list[float]] = {{col: [] for col in value_cols}}
    for row in rows:
        parsed = {{col: parse_float(row.get(col)) for col in value_cols}}
        if any(value is None for value in parsed.values()):
            continue
        label = str(row.get(spec["label_col"]) or len(labels) + 1)
        labels.append(label)
        for col in value_cols:
            series_values[col].append(float(parsed[col]))
        if len(labels) >= max_points:
            break
    readout_summary = summarize_series(labels, series_values)
    readability_warnings = chart_readability_warnings(
        labels,
        series_values,
        max_points=max_points,
        axis_label_min_pt=axis_label_min_pt,
    )
    analysis_metadata = {{
        "generated_by": "assets/make_figures.py",
        "producer_path": "assets/make_figures.py",
        "producer_sha256": producer_fingerprint["sha256"],
        "producer_bytes": producer_fingerprint["bytes"],
        "source_path": spec["source_path"],
        "source_label": spec.get("source_label", spec["source_path"]),
        "source_sha256": fingerprint["source_sha256"],
        "source_bytes": fingerprint["source_bytes"],
        "sheet_name": spec.get("sheet_name", ""),
        "label_col": spec["label_col"],
        "value_cols": value_cols,
        "selected_columns": selected_columns,
        "rows_scanned": len(rows),
        "rows_used": len(labels),
        "series_count": len(value_cols),
        "points": len(labels),
        "max_points": max_points,
        "chart_type": spec.get("chart_type", "bar"),
        "target_variant": spec.get("target_variant", ""),
        "target_box": spec.get("target_box", ""),
        "figure_size_inches": figure_size_inches,
        "figure_dpi": figure_dpi,
        "title_pt": title_pt,
        "axis_label_min_pt": axis_label_min_pt,
        "legend_pt": legend_pt,
        "x_label_rotation": x_label_rotation,
        "readout_summary": readout_summary,
        "readability_warnings": readability_warnings,
    }}

    chart_path = WORKSPACE / spec["chart_json"]
    figure_path = WORKSPACE / spec["figure_path"]
    table_path = WORKSPACE / spec["table_json"]
    series_payload = [
        {{"name": col, "values": series_values[col]}}
        for col in value_cols
    ]
    chart_payload = {{
        "name": spec["id"],
        "title": spec["title"],
        "type": spec.get("chart_type", "bar"),
        "categories": labels,
        "series": series_payload,
        "source_path": spec["source_path"],
        "source_label": spec.get("source_label", spec["source_path"]),
        "sheet_name": spec.get("sheet_name", ""),
        "provenance": f"Generated from {{spec.get('source_label', spec['source_path'])}} by assets/make_figures.py",
        "readability_warnings": readability_warnings,
        "analysis_metadata": analysis_metadata,
    }}

    table_rows: list[list[str]] = []
    summary_by_series = {{
        str(item.get("series")): item
        for item in readout_summary.get("series", [])
        if isinstance(item, dict)
    }}
    for col in value_cols:
        values = series_values[col]
        if values:
            mean_value = sum(values) / len(values)
            series_summary = summary_by_series.get(col, {{}})
            latest_text = format_number(values[-1])
            delta_value = series_summary.get("delta") if isinstance(series_summary, dict) else None
            delta_text = format_number(delta_value) if delta_value is not None else ""
            latest_delta = latest_text if not delta_text else "{{}} ({{}})".format(latest_text, delta_text)
            table_rows.append(
                [
                    col,
                    str(len(values)),
                    format_number(min(values)),
                    format_number(max(values)),
                    format_number(mean_value),
                    latest_delta,
                ]
            )
        else:
            table_rows.append([col, "0", "", "", "", ""])
    table_payload = {{
        "title": f"{{spec['title']}} summary",
        "headers": ["Series", "N", "Min", "Max", "Mean", "Latest (delta)"],
        "rows": table_rows,
        "caption": (
            f"Computed from {{spec.get('source_label', spec['source_path'])}} "
            f"by assets/make_figures.py; first {{max_points}} complete rows are used for slide readability."
        ),
        "source_path": spec["source_path"],
        "source_label": spec.get("source_label", spec["source_path"]),
        "sheet_name": spec.get("sheet_name", ""),
        "provenance": f"Summary table generated from {{spec.get('source_label', spec['source_path'])}} by assets/make_figures.py",
        "readout_summary": readout_summary,
        "readability_warnings": readability_warnings,
        "analysis_metadata": analysis_metadata,
    }}

    fig, ax = plt.subplots(figsize=tuple(figure_size_inches))
    colors = ["#0B6B78", "#C9302C", "#4F46E5", "#C77700"]
    chart_type = str(spec.get("chart_type", "bar")).lower()
    x_positions = list(range(len(labels)))
    if chart_type == "line":
        for idx, col in enumerate(value_cols):
            ax.plot(
                x_positions,
                series_values[col],
                marker="o",
                linewidth=2.0,
                markersize=4.0,
                color=colors[idx % len(colors)],
                label=col,
            )
        ax.set_xticks(x_positions)
        ax.set_xticklabels(labels, rotation=x_label_rotation, ha="right")
    elif len(value_cols) == 1:
        ax.bar(labels, series_values[value_cols[0]], color=colors[0])
    else:
        bar_width = min(0.28, 0.78 / max(1, len(value_cols)))
        offsets = [
            (idx - (len(value_cols) - 1) / 2.0) * bar_width
            for idx in range(len(value_cols))
        ]
        for idx, col in enumerate(value_cols):
            ax.bar(
                [x + offsets[idx] for x in x_positions],
                series_values[col],
                width=bar_width * 0.9,
                color=colors[idx % len(colors)],
                label=col,
            )
        ax.set_xticks(x_positions)
        ax.set_xticklabels(labels, rotation=x_label_rotation, ha="right")
    ax.set_title(spec["title"], fontsize=title_pt, loc="left")
    ax.set_ylabel(value_cols[0] if len(value_cols) == 1 else "Value")
    if len(value_cols) > 1:
        ax.legend(frameon=False, fontsize=legend_pt, loc="best")
    ax.tick_params(axis="x", labelrotation=x_label_rotation, labelsize=axis_label_min_pt)
    ax.tick_params(axis="y", labelsize=axis_label_min_pt)
    ax.grid(axis="y", alpha=0.22, linewidth=0.6)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    fig.tight_layout()
    save_slide_figure(fig, figure_path, dpi=figure_dpi)
    image_whitespace = image_exterior_whitespace(figure_path)
    analysis_metadata["image_whitespace"] = image_whitespace
    if image_whitespace.get("high_exterior_whitespace"):
        readability_warnings.append(
            "Generated figure has {{:.1f}}% exterior blank area; trim whitespace before final delivery.".format(
                float(image_whitespace.get("exterior_percent") or 0)
            )
        )
    analysis_metadata["readability_warnings"] = readability_warnings
    chart_payload["readability_warnings"] = readability_warnings
    table_payload["readability_warnings"] = readability_warnings
    write_json_if_changed(chart_path, chart_payload)
    write_json_if_changed(table_path, table_payload)
    return {{
        "id": spec["id"],
        "title": spec["title"],
        "chart_json": str(chart_path.relative_to(WORKSPACE)),
        "figure_path": str(figure_path.relative_to(WORKSPACE)),
        "table_json": str(table_path.relative_to(WORKSPACE)),
        "source_path": spec["source_path"],
        "source_label": spec.get("source_label", spec["source_path"]),
        "sheet_name": spec.get("sheet_name", ""),
        "label_col": spec["label_col"],
        "value_cols": value_cols,
        "selected_columns": selected_columns,
        "chart_type": spec.get("chart_type", "bar"),
        "series_count": len(value_cols),
        "points": len(labels),
        "readout_summary": readout_summary,
        "readability_warnings": readability_warnings,
        "analysis_metadata": analysis_metadata,
    }}


def build_artifact_manifest(outputs: list[dict[str, Any]]) -> dict[str, Any]:
    manifest_outputs: list[dict[str, Any]] = []
    for output in outputs:
        chart_id = str(output["id"])
        artifacts = [
            {{
                "id": f"{{chart_id}}_figure",
                "role": "figure",
                "alias": f"image:{{chart_id}}_figure",
                "path": output["figure_path"],
            }},
            {{
                "id": f"{{chart_id}}_chart_json",
                "role": "chart_json",
                "alias": f"chart:{{chart_id}}",
                "path": output["chart_json"],
            }},
            {{
                "id": f"{{chart_id}}_summary_table",
                "role": "summary_table",
                "alias": f"table:{{chart_id}}_summary",
                "path": output["table_json"],
            }},
        ]
        for artifact in artifacts:
            artifact["fingerprint"] = artifact_fingerprint(WORKSPACE / str(artifact["path"]))
        manifest_outputs.append(
            {{
                "id": chart_id,
                "title": output.get("title", ""),
                "source_path": output.get("source_path", ""),
                "source_label": output.get("source_label", output.get("source_path", "")),
                "series_count": output.get("series_count", 0),
                "points": output.get("points", 0),
                "selected_columns": output.get("selected_columns", []),
                "readout_summary": output.get("readout_summary", {{}}),
                "readability_warnings": output.get("readability_warnings", []),
                "analysis_metadata": output.get("analysis_metadata", {{}}),
                "artifacts": artifacts,
            }}
        )
    producer_fingerprint = artifact_fingerprint(Path(__file__))
    data_specs_hash = data_specs_sha256()
    rebuild_context = build_rebuild_context(
        outputs,
        manifest_outputs,
        producer_fingerprint=producer_fingerprint,
        data_specs_hash=data_specs_hash,
    )
    return {{
        "manifest_version": "presentation_skill_artifact_manifest_v1",
        "generated_by": "assets/make_figures.py",
        "producer_path": "assets/make_figures.py",
        "producer_sha256": producer_fingerprint["sha256"],
        "producer_bytes": producer_fingerprint["bytes"],
        "data_specs_sha256": data_specs_hash,
        "analysis_summary": "assets/analysis_summary.json",
        "analysis_summary_markdown": "assets/analysis_summary.md",
        "output_count": len(outputs),
        "rebuild_context": rebuild_context,
        "outputs": manifest_outputs,
    }}


def build_analysis_summary(outputs: list[dict[str, Any]], manifest: dict[str, Any]) -> dict[str, Any]:
    datasets: list[dict[str, Any]] = []
    source_paths: list[str] = []
    total_points = 0
    for output in outputs:
        metadata = output.get("analysis_metadata") if isinstance(output.get("analysis_metadata"), dict) else {{}}
        chart_id = str(output.get("id") or "")
        source_path = str(output.get("source_path") or metadata.get("source_path") or "")
        if source_path and source_path not in source_paths:
            source_paths.append(source_path)
        points = int(output.get("points") or metadata.get("points") or 0)
        total_points += points
        datasets.append(
            {{
                "id": chart_id,
                "title": str(output.get("title") or chart_id),
                "source_path": source_path,
                "source_label": str(output.get("source_label") or metadata.get("source_label") or source_path),
                "sheet_name": str(output.get("sheet_name") or metadata.get("sheet_name") or ""),
                "label_col": str(output.get("label_col") or metadata.get("label_col") or ""),
                "value_cols": list(output.get("value_cols") or metadata.get("value_cols") or []),
                "selected_columns": list(output.get("selected_columns") or metadata.get("selected_columns") or []),
                "chart_type": str(output.get("chart_type") or metadata.get("chart_type") or "bar"),
                "rows_scanned": int(metadata.get("rows_scanned") or 0),
                "rows_used": int(metadata.get("rows_used") or 0),
                "series_count": int(output.get("series_count") or metadata.get("series_count") or 0),
                "points": points,
                "readout_summary": output.get("readout_summary") or metadata.get("readout_summary") or {{}},
                "readability_warnings": list(output.get("readability_warnings") or metadata.get("readability_warnings") or []),
                "figure_path": str(output.get("figure_path") or ""),
                "chart_json": str(output.get("chart_json") or ""),
                "table_json": str(output.get("table_json") or ""),
                "aliases": {{
                    "figure": f"image:{{chart_id}}_figure",
                    "chart": f"chart:{{chart_id}}",
                    "table": f"table:{{chart_id}}_summary",
                }},
                "readability": {{
                    "target_box": str(metadata.get("target_box") or ""),
                    "figure_size_inches": metadata.get("figure_size_inches") or [],
                    "figure_dpi": metadata.get("figure_dpi"),
                    "axis_label_min_pt": metadata.get("axis_label_min_pt"),
                    "legend_pt": metadata.get("legend_pt"),
                    "x_label_rotation": metadata.get("x_label_rotation"),
                    "image_whitespace": metadata.get("image_whitespace") or {{}},
                }},
            }}
        )
    return {{
        "summary_version": "presentation_skill_analysis_summary_v1",
        "generated_by": "assets/make_figures.py",
        "producer_path": manifest.get("producer_path", "assets/make_figures.py"),
        "producer_sha256": manifest.get("producer_sha256", ""),
        "producer_bytes": manifest.get("producer_bytes", 0),
        "artifact_manifest": "assets/artifacts_manifest.json",
        "data_specs_sha256": manifest.get("data_specs_sha256", ""),
        "rebuild_context": manifest.get("rebuild_context") or {{}},
        "output_count": len(outputs),
        "source_paths": source_paths,
        "total_points": total_points,
        "datasets": datasets,
        "recommended_next_steps": [
            "Open this file before editing outline.json to choose the strongest generated evidence objects.",
            "Use aliases.figure for image-sidebar, aliases.chart for editable chart slides, or aliases.table for lab-run-results/table slides.",
            "After binding slides, keep artifact_registry.used_on_slides and figure_export_contract.target_slide aligned.",
        ],
    }}


def analysis_summary_markdown(summary: dict[str, Any]) -> str:
    rebuild_context = (
        summary.get("rebuild_context")
        if isinstance(summary.get("rebuild_context"), dict)
        else {{}}
    )
    rebuild_commands = (
        rebuild_context.get("commands")
        if isinstance(rebuild_context.get("commands"), dict)
        else {{}}
    )
    lines = [
        "# Generated Analysis Summary",
        "",
        f"- Outputs: {{summary.get('output_count', 0)}}",
        f"- Source files: {{', '.join(summary.get('source_paths') or []) or 'none'}}",
        f"- Total plotted points: {{summary.get('total_points', 0)}}",
        f"- Artifact manifest: {{summary.get('artifact_manifest', '')}}",
        f"- Producer: {{summary.get('producer_path', '')}} sha256 `{{summary.get('producer_sha256', '')}}`",
    ]
    if rebuild_context:
        lines.extend(
            [
                f"- Rebuild context: `{{rebuild_context.get('context_version', '')}}` sources `{{rebuild_context.get('source_count', 0)}}`",
                f"- Rebuild command: `{{rebuild_commands.get('rebuild_figures', '')}}`",
                f"- Inspect command: `{{rebuild_commands.get('inspect_manifest', '')}}`",
            ]
        )
    lines.extend(["", "## Datasets"])
    datasets = summary.get("datasets") if isinstance(summary.get("datasets"), list) else []
    for dataset in datasets:
        if not isinstance(dataset, dict):
            continue
        aliases = dataset.get("aliases") if isinstance(dataset.get("aliases"), dict) else {{}}
        readability = dataset.get("readability") if isinstance(dataset.get("readability"), dict) else {{}}
        whitespace = (
            readability.get("image_whitespace")
            if isinstance(readability.get("image_whitespace"), dict)
            else {{}}
        )
        lines.extend(
            [
                "",
                f"### {{dataset.get('title') or dataset.get('id')}}",
                f"- ID: `{{dataset.get('id', '')}}`",
                f"- Source: `{{dataset.get('source_path', '')}}`",
                f"- Columns: selected `{{', '.join(dataset.get('selected_columns') or [])}}`; label `{{dataset.get('label_col', '')}}`, values `{{', '.join(dataset.get('value_cols') or [])}}`",
                f"- Rows used: `{{dataset.get('rows_used', 0)}}` of `{{dataset.get('rows_scanned', 0)}}`; points `{{dataset.get('points', 0)}}`",
                f"- Primary readout: {{(dataset.get('readout_summary') or {{}}).get('primary') or 'none'}}",
                f"- Figure: `{{dataset.get('figure_path', '')}}` as `{{aliases.get('figure', '')}}`",
                f"- Chart JSON: `{{dataset.get('chart_json', '')}}` as `{{aliases.get('chart', '')}}`",
                f"- Table JSON: `{{dataset.get('table_json', '')}}` as `{{aliases.get('table', '')}}`",
                f"- Readability: target `{{readability.get('target_box', '')}}`, axis labels `{{readability.get('axis_label_min_pt', '')}}pt`",
            ]
        )
        if whitespace:
            if whitespace.get("checked"):
                lines.append(
                    f"- Figure whitespace: exterior `{{whitespace.get('exterior_percent', '')}}%`, content bbox `{{whitespace.get('content_bbox', [])}}`"
                )
            else:
                lines.append(
                    f"- Figure whitespace: not checked ({{whitespace.get('reason') or 'unknown reason'}})"
                )
        warnings = dataset.get("readability_warnings") if isinstance(dataset.get("readability_warnings"), list) else []
        if warnings:
            lines.append(f"- Readability notes: {{'; '.join(str(item) for item in warnings)}}")
    lines.extend(["", "## Next Steps"])
    for step in summary.get("recommended_next_steps") or []:
        lines.append(f"- {{step}}")
    return "\\n".join(lines).rstrip() + "\\n"


def main() -> int:
    outputs = [build_chart(spec) for spec in DATA_SPECS]
    manifest_path = WORKSPACE / "assets" / "artifacts_manifest.json"
    manifest = build_artifact_manifest(outputs)
    write_json_if_changed(manifest_path, manifest)
    analysis_summary_path = WORKSPACE / "assets" / "analysis_summary.json"
    analysis_summary_md_path = WORKSPACE / "assets" / "analysis_summary.md"
    summary = build_analysis_summary(outputs, manifest)
    write_json_if_changed(analysis_summary_path, summary)
    write_text_if_changed(analysis_summary_md_path, analysis_summary_markdown(summary))
    print(
        json.dumps(
            {{
                "outputs": outputs,
                "artifact_manifest": str(manifest_path.relative_to(WORKSPACE)),
                "analysis_summary": str(analysis_summary_path.relative_to(WORKSPACE)),
                "analysis_summary_markdown": str(analysis_summary_md_path.relative_to(WORKSPACE)),
            }},
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
'''


def _merge_unique(existing: list[Any], additions: list[Any]) -> list[Any]:
    seen: set[str] = set()
    merged: list[Any] = []
    for item in [*existing, *additions]:
        key = json.dumps(item, sort_keys=True, ensure_ascii=False) if isinstance(item, dict) else str(item)
        if key in seen:
            continue
        seen.add(key)
        merged.append(item)
    return merged


def _merge_named_entries(existing: list[Any], additions: list[dict[str, Any]]) -> list[Any]:
    merged: list[Any] = []
    positions: dict[str, int] = {}
    seen_raw: set[str] = set()
    for item in existing:
        if isinstance(item, dict) and str(item.get("name") or "").strip():
            key = _safe_id(str(item.get("name")))
            if key in positions:
                merged[positions[key]] = item
            else:
                positions[key] = len(merged)
                merged.append(item)
            continue
        raw_key = json.dumps(item, sort_keys=True, ensure_ascii=False) if isinstance(item, dict) else str(item)
        if raw_key in seen_raw:
            continue
        seen_raw.add(raw_key)
        merged.append(item)
    for item in additions:
        key = _safe_id(str(item.get("name") or ""))
        if key in positions:
            merged[positions[key]] = item
        else:
            positions[key] = len(merged)
            merged.append(item)
    return merged


def _artifact_registry_keys(item: dict[str, Any]) -> list[str]:
    keys: list[str] = []
    artifact_id = str(item.get("id") or "").strip()
    if artifact_id:
        keys.append(f"id:{artifact_id}")
    artifact_path = str(item.get("path") or "").strip().replace("\\", "/")
    if artifact_path:
        keys.append(f"path:{artifact_path}")
    return keys


def _merge_artifact_registry_entry(previous: dict[str, Any], current: dict[str, Any]) -> dict[str, Any]:
    merged = dict(previous)
    previous_slides = previous.get("used_on_slides")
    current_slides = current.get("used_on_slides")
    merged.update(current)
    merged["used_on_slides"] = _merge_unique(
        previous_slides if isinstance(previous_slides, list) else [],
        current_slides if isinstance(current_slides, list) else [],
    )
    for key in ("binding_status", "binding_note"):
        if key in previous and key not in current:
            merged[key] = previous[key]
    return merged


def _merge_artifact_registry(existing: list[Any], additions: list[dict[str, Any]]) -> list[Any]:
    merged: list[Any] = []
    positions: dict[str, int] = {}
    seen_raw: set[str] = set()

    def upsert(item: Any) -> None:
        if not isinstance(item, dict):
            raw_key = json.dumps(item, sort_keys=True, ensure_ascii=False) if isinstance(item, dict) else str(item)
            if raw_key in seen_raw:
                return
            seen_raw.add(raw_key)
            merged.append(item)
            return
        keys = _artifact_registry_keys(item)
        if not keys:
            raw_key = json.dumps(item, sort_keys=True, ensure_ascii=False)
            if raw_key in seen_raw:
                return
            seen_raw.add(raw_key)
            merged.append(item)
            return
        position = next((positions[key] for key in keys if key in positions), None)
        if position is None:
            position = len(merged)
            merged.append(item)
        else:
            previous = merged[position]
            merged[position] = (
                _merge_artifact_registry_entry(previous, item)
                if isinstance(previous, dict)
                else item
            )
        for key in _artifact_registry_keys(merged[position]) if isinstance(merged[position], dict) else keys:
            positions[key] = position

    for item in existing:
        upsert(item)
    for item in additions:
        upsert(item)
    return merged


def _data_specs_sha256(specs: list[dict[str, Any]]) -> str:
    return hashlib.sha256(
        json.dumps(specs, sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()


def _artifact_rebuild_context(
    specs: list[dict[str, Any]],
    *,
    script_rel: str,
    producer_sha256: str,
    producer_bytes: int,
) -> dict[str, Any]:
    source_paths = _merge_unique([], [str(spec.get("source_path") or "") for spec in specs])
    figure_paths = _merge_unique([], [str(spec.get("figure_path") or "") for spec in specs])
    chart_paths = _merge_unique([], [str(spec.get("chart_json") or "") for spec in specs])
    table_paths = _merge_unique([], [str(spec.get("table_json") or "") for spec in specs])
    return {
        "context_version": ARTIFACT_REBUILD_CONTEXT_VERSION,
        "generated_by": script_rel,
        "producer_path": script_rel,
        "producer_sha256": producer_sha256,
        "producer_bytes": producer_bytes,
        "data_specs_sha256": _data_specs_sha256(specs),
        "artifact_manifest": "assets/artifacts_manifest.json",
        "analysis_summary": "assets/analysis_summary.json",
        "analysis_summary_markdown": "assets/analysis_summary.md",
        "output_count": len(specs),
        "source_count": len(source_paths),
        "source_paths": source_paths,
        "artifact_paths": _merge_unique([], [*figure_paths, *chart_paths, *table_paths]),
        "outputs": {
            "figures": figure_paths,
            "chart_json": chart_paths,
            "summary_tables": table_paths,
        },
        "commands": {
            "rebuild_figures": f"python3 {script_rel}",
            "inspect_manifest": "python3 scripts/inspect_artifact_manifest.py --workspace <deck_workspace> --manifest assets/artifacts_manifest.json",
            "auto_select_lead": "python3 scripts/apply_artifact_manifest_bindings.py --workspace <deck_workspace> --manifest assets/artifacts_manifest.json --auto-select --auto-select-mode lead",
            "auto_select_all": "python3 scripts/apply_artifact_manifest_bindings.py --workspace <deck_workspace> --manifest assets/artifacts_manifest.json --auto-select",
            "validate_planning": "python3 scripts/validate_planning.py --workspace <deck_workspace>",
        },
        "command_cwd": {
            "rebuild_figures": "deck_workspace",
            "inspect_manifest": "presentation-skill repository",
            "auto_select_lead": "presentation-skill repository",
            "auto_select_all": "presentation-skill repository",
            "validate_planning": "presentation-skill repository",
        },
        "handoff_order": [
            "Run commands.rebuild_figures after editing DATA_SPECS or analysis code.",
            "Run commands.inspect_manifest to refresh alias, quality, and selection guidance.",
            "Run commands.auto_select_lead for a clean first pass or commands.auto_select_all for the full evidence triplet.",
            "Run commands.validate_planning before strict PPTX build.",
        ],
    }


def _update_design_brief(workspace: Path, specs: list[dict[str, Any]], script_rel: str) -> None:
    path = workspace / "design_brief.json"
    brief = _read_json(path, {})
    if not isinstance(brief, dict):
        brief = {}

    script_fingerprint = _file_fingerprint(workspace / script_rel)
    producer_sha256 = str(script_fingerprint.get("source_sha256") or "")
    producer_bytes = int(script_fingerprint.get("source_bytes") or 0)
    rebuild_context = _artifact_rebuild_context(
        specs,
        script_rel=script_rel,
        producer_sha256=producer_sha256,
        producer_bytes=producer_bytes,
    )

    outputs = [
        {
            "path": spec["figure_path"],
            "target_variant": spec["target_variant"],
            "target_box": spec["target_box"],
            "selected_columns": spec.get("selected_columns", []),
            "figure_size_inches": spec.get("figure_size_inches", DEFAULT_FIGURE_SIZE_INCHES),
            "figure_dpi": spec.get("figure_dpi", DEFAULT_FIGURE_DPI),
            "axis_label_min_pt": spec.get("axis_label_min_pt", DEFAULT_AXIS_LABEL_PT),
            "producer_path": script_rel,
            "producer_sha256": producer_sha256,
            "producer_bytes": producer_bytes,
            "title_pt": spec.get("title_pt", DEFAULT_TITLE_PT),
            "legend_pt": spec.get("legend_pt", DEFAULT_LEGEND_PT),
            "crop_rule": "Matplotlib bbox_inches='tight', pad_inches=0.04; run trim_image_whitespace.py if exterior whitespace remains",
            "readability_note": "6.4x3.6 in figure, 8pt axis labels minimum, one chart per slide region",
        }
        for spec in specs
    ]
    brief["figure_export_contract"] = {
        "script": script_rel,
        "script_sha256": producer_sha256,
        "script_bytes": producer_bytes,
        "rerun_command": f"python3 {script_rel}",
        "rebuild_context": rebuild_context,
        "outputs": outputs,
    }

    plan = brief.get("analysis_artifact_plan")
    if not isinstance(plan, dict):
        plan = {}
    data_files = [spec["source_path"] for spec in specs]
    plan["candidate_data_files"] = _merge_unique(list(plan.get("candidate_data_files", [])), data_files)
    plan["spreadsheet_inputs"] = _merge_unique(
        list(plan.get("spreadsheet_inputs", [])),
        [item for item in data_files if Path(item).suffix.lower() in {".xlsx", ".xls"}],
    )
    plan["figure_scripts"] = _merge_unique(list(plan.get("figure_scripts", [])), [script_rel])
    plan["artifact_manifest"] = "assets/artifacts_manifest.json"
    plan["analysis_summary"] = "assets/analysis_summary.json"
    plan["analysis_summary_markdown"] = "assets/analysis_summary.md"
    plan["rebuild_context"] = rebuild_context
    plan["chart_json_outputs"] = _merge_unique(list(plan.get("chart_json_outputs", [])), [spec["chart_json"] for spec in specs])
    plan["table_outputs"] = _merge_unique(list(plan.get("table_outputs", [])), [spec["table_json"] for spec in specs])
    plan["required_scripts"] = _merge_unique(list(plan.get("required_scripts", [])), [script_rel])
    plan["rebuild_commands"] = _merge_unique(list(plan.get("rebuild_commands", [])), [f"python3 {script_rel}"])
    registry_additions = []
    for spec in specs:
        source_label = str(spec.get("source_label") or spec["source_path"])
        value_cols = spec.get("value_cols") if isinstance(spec.get("value_cols"), list) else [spec["value_col"]]
        selected_columns = spec.get("selected_columns") if isinstance(spec.get("selected_columns"), list) else [spec["label_col"], *value_cols]
        value_label = ", ".join(str(item) for item in value_cols)
        source_path = workspace / str(spec["source_path"])
        fingerprint = _file_fingerprint(source_path)
        analysis_metadata = {
            "generated_by": script_rel,
            "producer_path": script_rel,
            "producer_sha256": producer_sha256,
            "producer_bytes": producer_bytes,
            "source_path": spec["source_path"],
            "source_label": source_label,
            "source_sha256": fingerprint.get("source_sha256", ""),
            "source_bytes": fingerprint.get("source_bytes", 0),
            "sheet_name": spec.get("sheet_name", ""),
            "label_col": spec["label_col"],
            "value_cols": value_cols,
            "selected_columns": selected_columns,
            "chart_type": spec.get("chart_type", "bar"),
            "max_points": MAX_POINTS,
            "target_variant": spec.get("target_variant", ""),
            "target_box": spec.get("target_box", ""),
            "figure_size_inches": spec.get("figure_size_inches", DEFAULT_FIGURE_SIZE_INCHES),
            "figure_dpi": spec.get("figure_dpi", DEFAULT_FIGURE_DPI),
            "title_pt": spec.get("title_pt", DEFAULT_TITLE_PT),
            "axis_label_min_pt": spec.get("axis_label_min_pt", DEFAULT_AXIS_LABEL_PT),
            "legend_pt": spec.get("legend_pt", DEFAULT_LEGEND_PT),
            "x_label_rotation": spec.get("x_label_rotation", DEFAULT_X_LABEL_ROTATION),
        }
        registry_additions.append(
            {
                "id": f"{spec['id']}_figure",
                "path": spec["figure_path"],
                "producer": script_rel,
                "used_on_slides": [],
                "provenance": f"Generated from {source_label} using label column {spec['label_col']} and value column(s) {value_label}",
                "analysis_metadata": {**analysis_metadata, "artifact_role": "figure"},
            }
        )
        registry_additions.append(
            {
                "id": f"{spec['id']}_chart_json",
                "path": spec["chart_json"],
                "producer": script_rel,
                "used_on_slides": [],
                "provenance": f"Editable chart spec generated from {source_label}",
                "analysis_metadata": {**analysis_metadata, "artifact_role": "chart_json"},
            }
        )
        registry_additions.append(
            {
                "id": f"{spec['id']}_summary_table",
                "path": spec["table_json"],
                "producer": script_rel,
                "used_on_slides": [],
                "provenance": f"Compact summary table generated from {source_label}",
                "analysis_metadata": {**analysis_metadata, "artifact_role": "summary_table"},
            }
        )
    plan["artifact_registry"] = _merge_artifact_registry(
        list(plan.get("artifact_registry", [])),
        registry_additions,
    )
    brief["analysis_artifact_plan"] = plan
    brief.setdefault(
        "readability_contract",
        {
            "min_title_pt": 24,
            "min_body_pt": 12,
            "min_caption_pt": 7.5,
            "max_title_lines": 2,
            "footer_reserved_inches": 0.34,
            "chart_label_min_pt": 8,
            "table_density_rule": "split or summarize tables that force unreadable text",
            "whitespace_rule": "avoid awkward empty regions; use figure/sidebar/table variants when content is sparse",
            "figure_crop_rule": "export tight bounding boxes and trim exterior whitespace before insertion",
        },
    )
    brief.setdefault(
        "speed_contract",
        {
            "renderer": "pptxgenjs by default; Python fallback only for legacy renderer-specific behavior",
            "first_pass": "run validate_planning.py and preflight before rendering",
            "render_policy": "render after data scripts and outline source are stable",
            "asset_policy": "reuse local generated figures and chart JSON before network assets",
            "conversion_hint": "use persistent LibreOffice/unoserver when available for repeated render QA",
        },
    )
    _write_json(path, brief)


def _update_asset_plan(workspace: Path, specs: list[dict[str, Any]]) -> None:
    path = workspace / "asset_plan.json"
    plan = _read_json(path, {"topic": workspace.name})
    if not isinstance(plan, dict):
        plan = {"topic": workspace.name}
    images = list(plan.get("images", [])) if isinstance(plan.get("images"), list) else []
    image_additions = [
        {
            "name": f"{spec['id']}_figure",
            "path": spec["figure_path"],
            "title": spec["title"],
            "source_path": spec["source_path"],
            "source_note": f"Generated from {spec.get('source_label') or spec['source_path']} by assets/make_figures.py",
            "provenance": "local generated analysis figure",
            "selected_columns": spec.get("selected_columns", []),
        }
        for spec in specs
    ]
    plan["images"] = _merge_named_entries(images, image_additions)
    charts = list(plan.get("charts", [])) if isinstance(plan.get("charts"), list) else []
    additions = [
        {
            "name": spec["id"],
            "path": spec["chart_json"],
            "title": spec["title"],
            "type": spec.get("chart_type", "bar"),
            "source_path": spec["source_path"],
            "source_label": spec.get("source_label") or spec["source_path"],
            "selected_columns": spec.get("selected_columns", []),
        }
        for spec in specs
    ]
    plan["charts"] = _merge_named_entries(charts, additions)
    tables = list(plan.get("tables", [])) if isinstance(plan.get("tables"), list) else []
    table_additions = [
        {
            "name": f"{spec['id']}_summary",
            "path": spec["table_json"],
            "title": f"{spec['title']} summary",
            "source_path": spec["source_path"],
            "source_label": spec.get("source_label") or spec["source_path"],
            "provenance": "local generated analysis summary table",
            "selected_columns": spec.get("selected_columns", []),
        }
        for spec in specs
    ]
    plan["tables"] = _merge_named_entries(tables, table_additions)
    plan.setdefault("backgrounds", [])
    plan.setdefault("generated_images", [])
    plan.setdefault("icons", [])
    _write_json(path, plan)


def _artifact_alias_plan(specs: list[dict[str, Any]], script_rel: str) -> list[dict[str, Any]]:
    plan: list[dict[str, Any]] = []
    for spec in specs:
        chart_id = str(spec["id"])
        image_alias = f"image:{chart_id}_figure"
        chart_alias = f"chart:{chart_id}"
        table_alias = f"table:{chart_id}_summary"
        source_path = str(spec["source_path"])
        source_label = str(spec.get("source_label") or source_path)
        selected_columns = list(spec.get("selected_columns") or [])
        common_sources = [source_path, script_rel]
        figure_caption = f"Generated figure from {source_label}."
        artifact_bindings = [
            {
                "artifact_id": f"{chart_id}_figure",
                "alias": image_alias,
                "path": spec["figure_path"],
                "role": "generated figure",
                "outline_field": "assets.hero_image",
                "used_on_slides_update": "append the chosen outline slide_id after inserting the alias",
            },
            {
                "artifact_id": f"{chart_id}_chart_json",
                "alias": chart_alias,
                "path": spec["chart_json"],
                "role": "editable native chart JSON",
                "outline_field": "assets.chart_data",
                "used_on_slides_update": "append the chosen outline slide_id after inserting the alias",
            },
            {
                "artifact_id": f"{chart_id}_summary_table",
                "alias": table_alias,
                "path": spec["table_json"],
                "role": "editable summary table JSON",
                "outline_field": "tables[] or table",
                "used_on_slides_update": "append the chosen outline slide_id after inserting the alias",
            },
        ]
        outline_field_snippets = [
            {
                "variant": "image-sidebar",
                "best_for": "dominant generated figure plus a concise interpretation sidebar",
                "fields": {
                    "variant": "image-sidebar",
                    "slide_intent": "evidence",
                    "visual_intent": "generated_figure",
                    "assets": {"hero_image": image_alias},
                    "caption": f"Generated from {source_label} by {script_rel}.",
                    "sources": common_sources,
                    "evidence_needs": [chart_id],
                    "required_artifact_ids": [f"{chart_id}_figure"],
                },
            },
            {
                "variant": "scientific-figure",
                "best_for": "figure-first report layout with compact caption and panel label",
                "fields": {
                    "variant": "scientific-figure",
                    "slide_intent": "evidence",
                    "visual_intent": "figure",
                    "figures": [
                        {
                            "path": image_alias,
                            "label": "A",
                            "caption": figure_caption,
                        }
                    ],
                    "caption": f"Generated from {source_label} by {script_rel}.",
                    "sources": common_sources,
                    "evidence_needs": [chart_id],
                    "required_artifact_ids": [f"{chart_id}_figure"],
                },
            },
            {
                "variant": "chart",
                "best_for": "editable native chart when the audience may revise labels or values in PowerPoint",
                "fields": {
                    "variant": "chart",
                    "slide_intent": "evidence",
                    "visual_intent": "data",
                    "assets": {"chart_data": chart_alias},
                    "sources": [source_path],
                    "evidence_needs": [chart_id],
                    "required_artifact_ids": [f"{chart_id}_chart_json"],
                },
            },
            {
                "variant": "lab-run-results",
                "best_for": "compact report table with summary statistics and source-line provenance",
                "fields": {
                    "variant": "lab-run-results",
                    "slide_intent": "evidence",
                    "visual_intent": "table",
                    "tables": [table_alias],
                    "sources": [spec["table_json"], source_path],
                    "evidence_needs": [chart_id],
                    "required_artifact_ids": [f"{chart_id}_summary_table"],
                },
            },
            {
                "variant": "table",
                "best_for": "single editable table when the preset prefers boardroom or decision-table treatment",
                "fields": {
                    "variant": "table",
                    "slide_intent": "evidence",
                    "visual_intent": "table",
                    "table_data": table_alias,
                    "sources": [spec["table_json"], source_path],
                    "evidence_needs": [chart_id],
                    "required_artifact_ids": [f"{chart_id}_summary_table"],
                },
            },
        ]
        plan.append(
            {
                "id": chart_id,
                "title": spec["title"],
                "source_path": source_path,
                "selected_columns": selected_columns,
                "image_alias": image_alias,
                "chart_alias": chart_alias,
                "table_alias": table_alias,
                "artifact_bindings": artifact_bindings,
                "outline_field_snippets": outline_field_snippets,
                "binding_updates": {
                    "analysis_artifact_plan": "after choosing slide ids, copy them into artifact_registry[*].used_on_slides for the matching artifact_id",
                    "figure_export_contract": "after choosing a figure slide, set outputs[*].target_slide and keep target_variant/target_box aligned with that slide",
                    "source_policy": "use compact source-line footers with short IDs; move long references to a final References slide",
                },
                "recommended_variants": [
                    {
                        "variant": "image-sidebar",
                        "use": "dominant generated figure plus interpretation sidebar",
                        "fields": outline_field_snippets[0]["fields"],
                    },
                    {
                        "variant": "chart",
                        "use": "editable native chart from generated chart JSON",
                        "fields": outline_field_snippets[1]["fields"],
                    },
                    {
                        "variant": "lab-run-results",
                        "use": "compact editable summary table from generated JSON",
                        "fields": outline_field_snippets[2]["fields"],
                    },
                ],
            }
        )
    return plan


def _artifact_handoff_payload(
    workspace: Path,
    manifest_path: Path,
    alias_plan: list[dict[str, Any]],
) -> dict[str, Any]:
    try:
        from inspect_artifact_manifest import _commands, _selection_template

        return {
            "selection_templates": [_selection_template(plan) for plan in alias_plan],
            "commands": _commands(workspace, manifest_path),
            "agent_next_steps": [
                "For a clean first pass, run commands.auto_select_lead; for the full figure/chart/table triplet, run commands.auto_select_all.",
                "For layout-guided ordering across all available variants, run commands.auto_select_recommended.",
                "For a custom subset, save one or more selection_templates[*].bindings entries to a selection file and run apply_artifact_manifest_bindings.py --selection.",
                "Revise generated titles, interpretation text, and sidebar notes in outline.json after binding.",
                "Run commands.validate_planning before commands.strict_build.",
            ],
        }
    except Exception as exc:  # pragma: no cover - report fallback only
        return {
            "selection_templates": [],
            "commands": {},
            "agent_next_steps": [],
            "handoff_payload_error": str(exc),
        }


def _collect_string_refs(value: Any) -> set[str]:
    refs: set[str] = set()

    def visit(item: Any) -> None:
        if isinstance(item, str):
            text = item.strip()
            if text:
                refs.add(text)
                refs.add(text.replace("\\", "/"))
                refs.add(text.lower())
        elif isinstance(item, list):
            for child in item:
                visit(child)
        elif isinstance(item, dict):
            for child in item.values():
                visit(child)

    visit(value)
    return refs


def _slide_ref(slide: dict[str, Any], idx: int) -> str:
    for key in ("slide_id", "id", "slug"):
        text = str(slide.get(key) or "").strip()
        if text:
            return text
    return f"s{idx}"


def _slide_variant(slide: dict[str, Any]) -> str:
    variant = str(slide.get("variant") or "").strip()
    if variant:
        return variant
    slide_type = str(slide.get("type") or "").strip()
    return slide_type if slide_type and slide_type != "content" else ""


def _slides_using_candidates(outline: Any, candidates: set[str]) -> list[dict[str, str]]:
    if not isinstance(outline, dict) or not isinstance(outline.get("slides"), list):
        return []
    normalized_candidates = {item for item in candidates if item}
    normalized_candidates.update(item.lower() for item in list(normalized_candidates))
    matches: list[dict[str, str]] = []
    for idx, slide in enumerate(outline.get("slides") or [], start=1):
        if not isinstance(slide, dict):
            continue
        refs = _collect_string_refs(slide)
        if refs.intersection(normalized_candidates):
            matches.append(
                {
                    "slide_ref": _slide_ref(slide, idx),
                    "variant": _slide_variant(slide),
                }
            )
    return matches


def _artifact_candidates(alias: str, path: str) -> set[str]:
    normalized_path = path.replace("\\", "/")
    return {
        alias,
        alias.lower(),
        normalized_path,
        normalized_path.lower(),
        Path(normalized_path).name,
        Path(normalized_path).name.lower(),
    }


def _merge_slide_refs(existing: Any, additions: list[str]) -> list[str]:
    base = [str(item).strip() for item in existing] if isinstance(existing, list) else []
    return _merge_unique([item for item in base if item], [item for item in additions if item])


def _bind_artifacts_to_outline(workspace: Path, specs: list[dict[str, Any]]) -> dict[str, Any]:
    outline = _read_json(workspace / "outline.json", None)
    if not isinstance(outline, dict) or not isinstance(outline.get("slides"), list):
        return {
            "applied": False,
            "reason": "outline.json missing or malformed",
            "bindings": [],
            "design_brief_changed": False,
            "asset_plan_changed": False,
        }

    bindings: list[dict[str, Any]] = []
    for spec in specs:
        chart_id = str(spec["id"])
        image_alias = f"image:{chart_id}_figure"
        chart_alias = f"chart:{chart_id}"
        table_alias = f"table:{chart_id}_summary"
        image_slides = _slides_using_candidates(
            outline,
            _artifact_candidates(image_alias, str(spec["figure_path"])),
        )
        chart_slides = _slides_using_candidates(
            outline,
            _artifact_candidates(chart_alias, str(spec["chart_json"])),
        )
        table_slides = _slides_using_candidates(
            outline,
            _artifact_candidates(table_alias, str(spec["table_json"])),
        )
        bindings.append(
            {
                "id": chart_id,
                "image_alias": image_alias,
                "chart_alias": chart_alias,
                "table_alias": table_alias,
                "image_slides": image_slides,
                "chart_slides": chart_slides,
                "table_slides": table_slides,
            }
        )

    brief_path = workspace / "design_brief.json"
    brief = _read_json(brief_path, {})
    if not isinstance(brief, dict):
        brief = {}
    plan = brief.get("analysis_artifact_plan")
    if not isinstance(plan, dict):
        plan = {}
    registry = plan.get("artifact_registry")
    if not isinstance(registry, list):
        registry = []
    registry_by_id = {
        str(item.get("id") or "").strip(): item
        for item in registry
        if isinstance(item, dict)
    }
    registry_by_path = {
        str(item.get("path") or "").strip().replace("\\", "/"): item
        for item in registry
        if isinstance(item, dict)
    }

    for spec, binding in zip(specs, bindings):
        artifact_specs = (
            (f"{spec['id']}_figure", spec["figure_path"], binding["image_slides"]),
            (f"{spec['id']}_chart_json", spec["chart_json"], binding["chart_slides"]),
            (f"{spec['id']}_summary_table", spec["table_json"], binding["table_slides"]),
        )
        for artifact_id, artifact_path, slide_matches in artifact_specs:
            slide_refs = [str(item.get("slide_ref") or "") for item in slide_matches if isinstance(item, dict)]
            if not slide_refs:
                continue
            entry = registry_by_id.get(str(artifact_id)) or registry_by_path.get(str(artifact_path).replace("\\", "/"))
            if isinstance(entry, dict):
                entry["used_on_slides"] = _merge_slide_refs(entry.get("used_on_slides"), slide_refs)

    plan["artifact_registry"] = registry
    brief["analysis_artifact_plan"] = plan

    figure_contract = brief.get("figure_export_contract")
    if isinstance(figure_contract, dict) and isinstance(figure_contract.get("outputs"), list):
        outputs = figure_contract.get("outputs") or []
        for output in outputs:
            if not isinstance(output, dict):
                continue
            output_path = str(output.get("path") or "").strip().replace("\\", "/")
            for spec, binding in zip(specs, bindings):
                if output_path != str(spec["figure_path"]).replace("\\", "/"):
                    continue
                image_slides = [item for item in binding.get("image_slides", []) if isinstance(item, dict)]
                if not image_slides:
                    continue
                first_slide = image_slides[0]
                output["target_slide"] = first_slide.get("slide_ref") or output.get("target_slide") or ""
                variant = str(first_slide.get("variant") or "").strip()
                if variant:
                    output["target_variant"] = variant
                break

    asset_path = workspace / "asset_plan.json"
    asset_plan = _read_json(asset_path, {})
    if not isinstance(asset_plan, dict):
        asset_plan = {}
    asset_sections = {
        "images": lambda spec, binding: (f"{spec['id']}_figure", binding["image_slides"]),
        "charts": lambda spec, binding: (str(spec["id"]), binding["chart_slides"]),
        "tables": lambda spec, binding: (f"{spec['id']}_summary", binding["table_slides"]),
    }
    for section, selector in asset_sections.items():
        entries = asset_plan.get(section)
        if not isinstance(entries, list):
            continue
        for spec, binding in zip(specs, bindings):
            expected_name, slide_matches = selector(spec, binding)
            slide_refs = [str(item.get("slide_ref") or "") for item in slide_matches if isinstance(item, dict)]
            if not slide_refs:
                continue
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                if str(entry.get("name") or "").strip() == expected_name:
                    entry["used_on_slides"] = _merge_slide_refs(entry.get("used_on_slides"), slide_refs)

    design_changed = _write_json(brief_path, brief)
    asset_changed = _write_json(asset_path, asset_plan)
    return {
        "applied": True,
        "reason": "",
        "bindings": bindings,
        "design_brief_changed": design_changed,
        "asset_plan_changed": asset_changed,
    }


def _args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scaffold reproducible figure, chart, and summary-table artifacts from local tabular data.")
    parser.add_argument("--workspace", required=True, help="Deck workspace directory")
    parser.add_argument(
        "--data-path",
        action="append",
        default=[],
        help="Data file or directory. Defaults to workspace data/assets data locations.",
    )
    parser.add_argument(
        "--script-path",
        default="assets/make_figures.py",
        help="Workspace-relative figure script path to write.",
    )
    parser.add_argument("--report", help="Optional JSON report path")
    parser.add_argument("--no-apply", action="store_true", help="Do not update design_brief.json or asset_plan.json")
    parser.add_argument(
        "--bind-outline",
        action="store_true",
        help=(
            "Inspect outline.json for generated artifact aliases/paths and update "
            "used_on_slides plus figure_export_contract target_slide fields."
        ),
    )
    parser.add_argument("--run", action="store_true", help="Run the generated figure script after writing it")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite an existing figure script")
    return parser.parse_args()


def main() -> int:
    args = _args()
    workspace = Path(args.workspace).expanduser().resolve()
    if not workspace.exists() or not workspace.is_dir():
        print(f"Error: workspace not found: {workspace}", file=sys.stderr)
        return 1

    candidates = _candidate_data_files(workspace, args.data_path)
    specs: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    for path in candidates:
        if not path.exists():
            skipped.append({"path": str(path), "reason": "missing"})
            continue
        try:
            inferred_specs, inferred_skipped = _infer_chart_specs(workspace, path)
        except Exception as exc:
            skipped.append({"path": str(path), "reason": str(exc)})
            continue
        specs.extend(inferred_specs)
        skipped.extend(inferred_skipped)
    _disambiguate_spec_ids(specs)

    if not specs:
        report = {"workspace": str(workspace), "specs": [], "alias_plan": [], "skipped": skipped, "updated": False}
        if args.report:
            _write_json(Path(args.report).expanduser().resolve(), report)
        print(json.dumps(report, indent=2))
        return 1

    script_rel = str(Path(args.script_path))
    script_path = workspace / script_rel
    script_text = _make_figures_script(specs)
    script_updated = True
    script_unchanged = False
    if script_path.exists():
        try:
            existing_script = script_path.read_text(encoding="utf-8")
        except OSError:
            existing_script = ""
        if existing_script == script_text:
            script_updated = False
            script_unchanged = True
        elif not args.overwrite:
            print(
                f"Error: figure script exists and differs; pass --overwrite to replace: {script_path}",
                file=sys.stderr,
            )
            return 2
    script_path.parent.mkdir(parents=True, exist_ok=True)
    if script_updated:
        script_path.write_text(script_text, encoding="utf-8")
    script_path.chmod(0o755)

    if not args.no_apply:
        _update_design_brief(workspace, specs, script_rel)
        _update_asset_plan(workspace, specs)

    outline_binding: dict[str, Any] = {
        "applied": False,
        "reason": "--bind-outline not requested",
        "bindings": [],
        "design_brief_changed": False,
        "asset_plan_changed": False,
    }
    if args.bind_outline:
        if args.no_apply:
            outline_binding = {
                "applied": False,
                "reason": "--no-apply prevents outline binding writes",
                "bindings": [],
                "design_brief_changed": False,
                "asset_plan_changed": False,
            }
        else:
            outline_binding = _bind_artifacts_to_outline(workspace, specs)

    run_result: dict[str, Any] = {"ran": False}
    if args.run:
        result = subprocess.run(
            [sys.executable, str(script_path)],
            cwd=str(workspace),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )
        run_result = {
            "ran": True,
            "returncode": result.returncode,
            "stdout": result.stdout[-4000:],
        }
        if result.returncode != 0:
            print(result.stdout, end="")
            return result.returncode

    alias_plan = _artifact_alias_plan(specs, script_rel)
    manifest_inspection_error = ""
    manifest_path = workspace / "assets" / "artifacts_manifest.json"
    handoff_payload: dict[str, Any] = {}
    if args.run and manifest_path.exists():
        try:
            from inspect_artifact_manifest import inspect_manifest

            inspected = inspect_manifest(workspace, manifest_path)
            inspected_alias_plan = inspected.get("alias_plan")
            if isinstance(inspected_alias_plan, list) and inspected_alias_plan:
                alias_plan = inspected_alias_plan
            handoff_payload = {
                key: inspected.get(key)
                for key in ("selection_templates", "commands", "agent_next_steps")
                if key in inspected
            }
        except Exception as exc:  # pragma: no cover - report fallback only
            manifest_inspection_error = str(exc)
    if not handoff_payload:
        handoff_payload = _artifact_handoff_payload(workspace, manifest_path, alias_plan)
    script_fingerprint = _file_fingerprint(script_path)
    rebuild_context = _artifact_rebuild_context(
        specs,
        script_rel=script_rel,
        producer_sha256=str(script_fingerprint.get("source_sha256") or ""),
        producer_bytes=int(script_fingerprint.get("source_bytes") or 0),
    )

    report = {
        "workspace": str(workspace),
        "script": str(script_path),
        "script_rel": script_rel,
        "artifact_manifest": "assets/artifacts_manifest.json",
        "analysis_summary": "assets/analysis_summary.json",
        "analysis_summary_markdown": "assets/analysis_summary.md",
        "rebuild_context": rebuild_context,
        "specs": specs,
        "alias_plan": alias_plan,
        **handoff_payload,
        "manifest_inspection_error": manifest_inspection_error,
        "outline_binding": outline_binding,
        "skipped": skipped,
        "updated": not args.no_apply,
        "script_updated": script_updated,
        "script_unchanged": script_unchanged,
        "run": run_result,
    }
    if args.report:
        _write_json(Path(args.report).expanduser().resolve(), report)
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
